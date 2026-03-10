import streamlit as st
import pandas as pd
import io
import zipfile
import json
import os
from datetime import datetime
from openpyxl import Workbook

# =========================================================================
# 1. 初始化配置与数据加载
# =========================================================================

def load_configurations():
    """读取配置文件并加载CSV模版数据"""
    config_path = 'config.json'
    
    # 检查配置文件是否存在
    if not os.path.exists(config_path):
        st.error(f"❌ 配置文件丢失: {config_path}")
        return {}

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            raw_config = json.load(f)
        
        # 遍历配置，加载对应的 CSV 文件
        for channel, data in raw_config.items():
            csv_file = data.get("file_path")
            if os.path.exists(csv_file):
                # 读取 CSV，无表头模式
                df = pd.read_csv(csv_file, header=None)
                # 为 DataFrame 赋予列名（用于后续处理或调试，虽然后续逻辑主要是按行处理）
                if "columns" in data and len(data["columns"]) == len(df.columns):
                    df.columns = data["columns"]
                
                # 将数据转为字典列表存储在 rows 字段中，保持原有逻辑兼容
                data["rows"] = df.to_dict(orient="records")
            else:
                st.warning(f"⚠️ 渠道 [{channel}] 的模版文件未找到: {csv_file}")
                data["rows"] = []
                
        return raw_config
    except Exception as e:
        st.error(f"加载配置失败: {e}")
        return {}

# 加载数据 (使用 st.cache_data 避免每次刷新都重读 IO)
@st.cache_data
def get_raw_data():
    return load_configurations()

RAW_DATA = get_raw_data()

# =========================================================================
# 2. 核心处理逻辑
# =========================================================================

def clean_val(val):
    if pd.isna(val) or val == "":
        return ""
    s_val = str(val)
    # 清理零宽字符 (Zero-Width Space, Zero-Width Non-Joiner 等)
    s_val = s_val.replace('\u200b', '').replace('\u200c', '').replace('\u200d', '').replace('\ufeff', '')
    if s_val.endswith(".0"):
        return s_val[:-2]
    return s_val

def process_rows(channel, p_name, p_id, t_id):
    """根据用户输入，深度复制模板行并替换关键字段"""
    template_conf = RAW_DATA.get(channel)
    if not template_conf or not template_conf.get("rows"):
        return pd.DataFrame()

    rows = template_conf["rows"]
    sample_name = template_conf["sample_name"]
    sample_pid = template_conf["sample_pid"]
    sample_adid = template_conf["sample_adid"]
    columns = template_conf["columns"]
    
    new_rows = []
    for r in rows:
        new_row = r.copy()
        
        # 遍历该行的所有列进行替换
        for k, v in new_row.items():
            val_str = clean_val(v)
            
            # 替换逻辑
            if sample_pid in val_str:
                val_str = val_str.replace(sample_pid, str(p_id))
            if sample_adid in val_str:
                val_str = val_str.replace(sample_adid, str(t_id))
            if sample_name in val_str:
                val_str = val_str.replace(sample_name, p_name)
                
            new_row[k] = val_str
            
        new_rows.append(new_row)
        
    # 确保列顺序一致
    return pd.DataFrame(new_rows, columns=columns)

def create_xlsx_file(df):
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'
    
    # 写入表头
    for col_idx, header in enumerate(df.columns, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)
    
    # 写入数据行
    for row_idx, row_data in df.iterrows():
        for col_idx, value in enumerate(row_data, start=1):
            val = value
            # 尝试将字符串转换为数字，解决 Excel "文本存储为数字" 的问题
            if isinstance(val, str):
                val_stripped = val.strip()
                try:
                    f_val = float(val_stripped)
                    if f_val.is_integer():
                        val = int(f_val)
                    else:
                        val = f_val
                except (ValueError, TypeError):
                    pass
            worksheet.cell(row=row_idx + 2, column=col_idx, value=val)
            
    workbook.save(output)
    output.seek(0)
    return output

# =========================================================================
# 3. Streamlit 界面
# =========================================================================
st.set_page_config(page_title="广告配置生成器", layout="wide")
# =========================================================================
# 👇👇👇 在这里添加隐藏代码 👇👇👇
# =========================================================================
hide_streamlit_style = """
    <style>
    .stDeployButton {
        visibility: hidden;
    }
    footer {
        visibility: hidden;
    }
    header {
        visibility: hidden;
    }
    /* 隐藏右下角 Streamlit Community Cloud 悬浮图标 */
    ._profileContainer_gzau3_53 {
        visibility: hidden;
    }
    [data-testid="manage-app-button"] {
        visibility: hidden;
    }
    .viewerBadge_container__r5tak {
        visibility: hidden;
    }
    div[class*="StatusWidget"] {
        visibility: hidden;
    }
    iframe[title="chat"] {
        display: none;
    }
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)
# =========================================================================
st.title("广告配置自动化工具")

# 侧边栏说明
# 侧边栏说明
with st.sidebar:
    # 版本信息
    st.markdown("### 📋 工具信息")
    st.markdown("**版本:** v1.2.0 · 2026-03-10")
    with st.expander("📝 更新日志"):
        st.markdown("""
**v1.2.0** (2026-03-10)
- 新增穿山甲new50/优量汇new50渠道模板
- 移除旧版穿山甲new1/优量汇new2
- 隐藏右下角 Streamlit 悬浮图标

**v1.1.0** (2026-03-06)
- 修复穿山甲new1/优量汇new2配置替换失效
- 输出格式从 .xls 升级为 .xlsx
- 新增输入校验（非空/纯数字检查）
- 修复 CSV 模板零宽字符问题

**v1.0.0** (初始版本)
- 支持穿山甲/优量汇/快手渠道
- 批量生成广告配置文件
        """)
    
    st.markdown("---")
    
    # 使用说明
    with st.expander("📖 使用说明"):
        st.markdown("""
1. **选择渠道** — 勾选需要生成配置的广告平台
2. **填写产品信息** — 在表格中输入产品ID、名称、广告ID（均为纯数字）
3. **批量生成** — 点击生成按钮，下载包含所有配置的 zip 包
4. 支持多行输入，一次生成多个产品的配置

⚠️ 产品ID和广告ID请从对应平台后台获取
        """)
    
    st.markdown("---")
    
    # 模板下载
    with st.expander("📁 模板 CSV 下载"):
        if RAW_DATA:
            for ch, data in RAW_DATA.items():
                csv_path = data.get("file_path", "")
                if os.path.exists(csv_path):
                    with open(csv_path, 'rb') as f:
                        st.download_button(
                            label=f"⬇️ {ch}",
                            data=f.read(),
                            file_name=os.path.basename(csv_path),
                            mime="text/csv",
                            key=f"dl_{ch}"
                        )
                else:
                    st.caption(f"{ch}: 文件不存在")
    
    st.markdown("---")
    
    # 配置加载状态
    st.info("配置加载状态：")
    if RAW_DATA:
        for ch in RAW_DATA.keys():
            row_count = len(RAW_DATA[ch].get('rows', []))
            note = RAW_DATA[ch].get('note', '')
            label = f"{ch} ({note})" if note else ch
            st.success(f"✅ {label}: {row_count} 条模版")
    else:
        st.error("❌ 未加载到任何配置，请检查 json 文件。")

# 1. 选择渠道 (动态从配置中读取 key)
st.subheader("1. 选择广告渠道")
available_channels = list(RAW_DATA.keys()) if RAW_DATA else []
selected_channels = st.multiselect(
    "请选择目标平台:", 
    options=available_channels, 
    default=[available_channels[0]] if available_channels else None
)

# 2. 输入表格
st.subheader("2. 批量输入产品信息")

if 'input_df' not in st.session_state:
    st.session_state.input_df = pd.DataFrame([{
        "产品ID": "12345", 
        "产品名称": "示例产品", 
        "广告ID": "67890"
    }])

edited_df = st.data_editor(
    st.session_state.input_df,
    num_rows="dynamic",
    column_config={
        "产品ID": st.column_config.TextColumn("产品ID", help="对应平台的应用ID"),
        "产品名称": st.column_config.TextColumn("产品名称", help="新产品的中文名称"),
        "广告ID": st.column_config.TextColumn("广告ID", help="对应平台的广告位ID (tappid)"),
    }
)

# 3. 生成
st.markdown("---")
if st.button("🚀 立即生成配置文档", type="primary"):
    if not selected_channels:
        st.error("请至少选择一个渠道！")
    elif edited_df.empty:
        st.error("请填写产品信息！")
    else:
        # 输入校验
        has_error = False
        for idx, row in edited_df.iterrows():
            p_name = str(row.get("产品名称", "")).strip()
            p_id = str(row.get("产品ID", "")).strip()
            t_id = str(row.get("广告ID", "")).strip()
            
            if not p_name or p_name == "nan":
                st.error(f"第 {idx+1} 行：产品名称不能为空")
                has_error = True
            if not p_id or p_id == "nan":
                st.error(f"第 {idx+1} 行：产品ID不能为空")
                has_error = True
            elif not p_id.isdigit():
                st.error(f"第 {idx+1} 行：产品ID必须为纯数字，当前值: {p_id}")
                has_error = True
            if not t_id or t_id == "nan":
                st.error(f"第 {idx+1} 行：广告ID不能为空")
                has_error = True
            elif not t_id.isdigit():
                st.error(f"第 {idx+1} 行：广告ID必须为纯数字，当前值: {t_id}")
                has_error = True
        
        if has_error:
            st.warning("请修正以上错误后再生成。")
        else:
            zip_buffer = io.BytesIO()
            file_count = 0
            
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for idx, row in edited_df.iterrows():
                    p_name = row.get("产品名称", "")
                    p_id = str(row.get("产品ID", ""))
                    t_id = str(row.get("广告ID", ""))
                    
                    if not p_name or not p_id: 
                        continue
                    
                    for ch in selected_channels:
                        final_df = process_rows(ch, p_name, p_id, t_id)
                        
                        if not final_df.empty:
                            xls_data = create_xlsx_file(final_df)
                            fname = f"{ch}_{p_name}.xlsx"
                            zf.writestr(fname, xls_data.getvalue())
                            file_count += 1
            
            if file_count > 0:
                st.success(f"✅ 成功生成 {file_count} 个文件！")
                st.download_button(
                    label="📥 下载所有文件 (.zip)",
                    data=zip_buffer.getvalue(),
                    file_name=f"广告配置_{datetime.now().strftime('%H%M%S')}.zip",
                    mime="application/zip"
                )
            else:
                st.warning("未生成任何文件，请检查输入数据或模版配置。")